import copy
import csv
import logging
import os
import re
import stat
import sys
import threading
import time
from datetime import datetime
from inspect import currentframe
from threading import Thread

import cv2
import matplotlib
import numpy as np
import openpyxl
import psutil
import pyautogui
import zxing
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import Qt, QRegExp, QMetaObject, QTimer, QUrl, QObject, QEvent, QRect
from PyQt5.QtGui import QIcon, QCursor, QBrush, QRegExpValidator, QPixmap, QImage, QDesktopServices, \
    QStandardItemModel, QStandardItem, QColor
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtWidgets import QMessageBox, QMenu, QApplication, QAbstractItemView, QHeaderView, QLabel, QAction, \
    QInputDialog, QLineEdit, QToolTip
from matplotlib import pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import conf.config
import conf.globalvar as gv
import dal.database.sqlite
import models.loadseq
import models.product
import models.testcase
import models.variables
from bll.testthread import TestThread, TestStatus
from common.basicfunc import IsNullOrEmpty, run_cmd, create_csv_file, GetAllIpv4Address, str_to_int, ensure_path_sep
from common.mysignals import update_label, on_setIcon, updateAction, controlEnable, on_actionLogFolder, \
    on_actionException
from common.testform import TestForm
from conf.logprint import QTextEditHandler, LogPrint
from peak.plin.peaklin import PeakLin
from ui.settings import SettingsDialog
from ui.ui_main import Ui_MainWindow

matplotlib.use("Qt5Agg")


# pyrcc5 images.qrc -o images.py
# pyuic5 ui_main.ui -o main_ui.py

def column_width_autofit(ws):
    # 设置一个字典用于保存列宽数据
    dims = {}
    # 遍历表格数据，获取自适应列宽数据
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        ws.column_dimensions[get_column_letter(col)].width = value + 2


class MainForm(Ui_MainWindow, TestForm):
    main_form = None
    _lock = threading.RLock()

    def __new__(cls, *args, **kwargs):

        if cls.main_form:  # 如果已经有单例了就不再去抢锁，避免IO等待
            return cls.main_form
        with cls._lock:  # 使用with语法，方便抢锁释放锁
            if not cls.main_form:
                cls.main_form = super().__new__(cls, *args, **kwargs)
            return cls.main_form

    def __init__(self):
        Ui_MainWindow.__init__(self)
        TestForm.__init__(self)
        self.treeViewModel = None
        self.tableViewRetModel = None
        self.model = None
        self.setupUi(self)
        self.stepClipboard = None
        self.suitClipboard = None
        self.header_new = []
        self.graphic_scene = None
        self.canvas = None
        self.t = 0
        self.QtimerID = None
        self.tableWidgetHeader = ["SN", "StepName", "SPEC", "LSL", "Value", "USL", "Time", "StartTime", "Result"]
        self.setWindowTitle(self.windowTitle() + f' v{gv.VERSION}')
        gv.InitCreateDirs(self.logger)
        self.init_textEditHandler()
        self.init_lab_factory(gv.cfg.station.privileges)
        self.init_ViewWidget()
        self.init_childLabel()
        self.init_label_info()
        self.init_lineEdit()
        self.init_graphicsView()
        self.init_signals_connect()
        self.testcase: models.testcase.TestCase = models.testcase.TestCase(rf'{gv.ExcelFilePath}',
                                                                           f'{gv.cfg.station.station_name}',
                                                                           self.logger,
                                                                           self)
        self.init_status_bar()
        self.init_select_station()
        self.testSequences = self.testcase.cloneSuites
        self.ShowTreeView(self.testSequences)
        self.testThread = TestThread(self)
        self.testThread.start()
        self.open_camera()
        self.closeEvent = self.closeEvent

    def init_select_station(self):
        """
        #    - FL_M4 #(前左):0x1E
        #    - FR_M4 #(前右):0x50
        #    - RL_M4 #(后左):0x1F
        #    - RML_M4 #(后中左):0x51
        #    - RMR_M4 #(后中右):0x52
        #    - RR_M4 #(后右):0x53
        #    - FL_SX5GEV #(前左):0x1E
        #    - FR_SX5GEV #(前右):0x50
        #    - RL_SX5GEV #(后左):0x1F
        #    - RM_SX5GEV #(后中):0x51
        #    - RR_SX5GEV #(后右):0x53
        #    - FLB_SX5GEV #(前左保杠):0x52
        #    - FRB_SX5GEV #(前右保杠):0x54
        #    - FL_M6 #(前左):0x1E
        #    - FLG_M6 #(左格栅灯):0x52
        #    - FR_M6 #(前右):0x50
        #    - FRG_M6 #(右格栅灯):0x54
        """
        station_model = ['M4', 'M6', 'SX5GEV', 'SX5GEV_EOL']
        flag = True
        for item in gv.cfg.station.station_all:
            model_all = []
            if item not in station_model:
                station_select = QAction(self)
                station_select.setObjectName(item)
                station_select.setText(item)
                self.menuSelect_Station.addAction(station_select)
                station_select.triggered.connect(self.on_selectStation)
            else:
                if item == 'M4':
                    model_all = ['FL', 'FR', 'RL', 'RML', 'RMR', 'RR']
                elif item == 'M6':
                    model_all = ['FL', 'FR', 'FLG', 'FRG']
                elif item == 'SX5GEV' or item == 'SX5GEV_EOL':
                    model_all = ['FL', 'FR', 'RL', 'RM', 'RR', 'FLB', 'FRB']

                station_menu = QMenu(self.menuSelect_Station)
                station_menu.setObjectName(item)
                station_menu.setTitle(item)
                self.menuSelect_Station.addMenu(station_menu)
                station_menu.triggered.connect(self.on_selectStation)
                for item_ in model_all:
                    model_select = QAction(self)
                    model_select.setObjectName(item_)
                    model_select.setText(item_)
                    station_menu.addAction(model_select)
                    model_select.triggered.connect(self.on_selectDutModel)
                    if gv.cfg.station.station_name == station_menu.title() and flag:
                        flag = False
                        model_select.triggered.emit()

    def init_textEditHandler(self):
        """create log handler for textEdit"""
        textEdit_handler = QTextEditHandler(stream=self.textEdit)
        textEdit_handler.name = 'textEdit_handler'
        log_console = logging.getLogger('testlog').handlers[0]
        if getattr(sys, 'frozen', False):
            logging.getLogger('testlog').removeHandler(log_console)
            textEdit_handler.formatter = gv.lg.logger.handlers[0].formatter
            if gv.IsHide:
                textEdit_handler.level = gv.lg.logger.handlers[1].level
                self.fileHandle = gv.lg.logger.handlers[2]
            else:
                textEdit_handler.level = gv.lg.logger.handlers[0].level
                self.fileHandle = gv.lg.logger.handlers[0]
        else:
            gv.cfg.station.privileges = 'lab'
            textEdit_handler.formatter = gv.lg.logger.handlers[1].formatter
            textEdit_handler.level = gv.lg.logger.handlers[1].level
            self.fileHandle = gv.lg.logger.handlers[1]
        logging.getLogger('testlog').addHandler(textEdit_handler)

    def init_lineEdit(self):
        self.lineEdit.setFocus()
        self.lineEdit.setMaxLength(gv.cfg.dut.sn_len)
        # 自定义文本验证器
        reg = QRegExp('^[A-Z0-9]{' + f'{gv.cfg.dut.sn_len},' + f'{gv.cfg.dut.sn_len}' + '}')
        pValidator = QRegExpValidator(self.lineEdit)
        pValidator.setRegExp(reg)
        # if not gv.IsDebug:
        #     self.lineEdit.setValidator(pValidator)

    def init_graphicsView(self):
        self.canvas = FigureCanvas(self.sinWave())
        self.graphic_scene = QtWidgets.QGraphicsScene()
        # print(self.graphicsView.geometry())
        x = self.graphicsView.x()
        y = self.graphicsView.y()
        self.graphic_scene.setSceneRect(x, y, self.graphicsView.width() * 0.98, self.graphicsView.height() * 0.9)
        # print(self.graphic_scene.width(), self.graphic_scene.height(), self.graphic_scene.sceneRect())
        self.graphic_scene.addWidget(self.canvas)
        self.graphicsView.setScene(self.graphic_scene)

    def init_childLabel(self):
        self.lb_failInfo = QLabel('Next:O-SFT /Current:O', self.lb_status)
        self.lb_failInfo.setStyleSheet(f"background-color:#f0f0f0;font: 11pt '宋体';")
        self.lb_failInfo.setHidden(True)
        self.lb_testTime = QLabel('TestTime:30s', self.lb_errorCode)
        self.lb_testTime.setStyleSheet(f"background-color:#f0f0f0;font: 11pt '宋体';")
        self.lb_testTime.setHidden(True)

    def init_ViewWidget(self):
        strHeaderQss = "QHeaderView::section { background:#CCCCCC; color:black;min-height:2em;}"
        if not gv.IsDebug:
            self.tableViewStepProp.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            self.tableViewVar.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableViewStepProp.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableViewStepProp.horizontalHeader().setStyleSheet(strHeaderQss)
        self.tableViewStepProp.installEventFilter(self)
        self.stepMenu = QMenu(self.tableViewStepProp)
        self.stepMenu.addAction('Insert Row')
        self.stepMenu.addAction('Delete Row')
        self.stepMenu.actions()[0].triggered.connect(self.on_stepInsertRow)
        self.stepMenu.actions()[1].triggered.connect(self.on_stepDeleteRow)

        self.tableViewVar.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableViewVar.horizontalHeader().setStyleSheet(strHeaderQss)

        self.tableViewRetModel = QStandardItemModel(0, 9, self.tableViewRet)
        self.tableViewRetModel.setHorizontalHeaderLabels(self.tableWidgetHeader)
        self.tableViewRet.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableViewRet.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableViewRet.horizontalHeader().setStyleSheet(strHeaderQss)
        self.tableViewRet.setModel(self.tableViewRetModel)
        self.tableViewRet.resizeColumnsToContents()
        self.tableViewRet.resizeRowsToContents()

        self.treeViewModel = QStandardItemModel(0, 1, self.treeView)

    def init_lab_factory(self, str_):
        if str_ == "lab":
            gv.IsDebug = True
            self.actionPrivileges.setIcon(QIcon(':/images/UnlockedObject.ico'))
        else:
            gv.IsDebug = False
            self.actionPrivileges.setIcon(QIcon(':/images/LockedObject.ico'))
            self.actionConvertExcelToJson.setEnabled(False)
            self.actionReloadScript.setEnabled(False)
            self.actionStart.setEnabled(False)
            self.actionStop.setEnabled(False)
            self.actionClearLog.setEnabled(False)
            if getattr(sys, 'frozen', False):
                self.actionStart.setEnabled(True)
                self.actionConfig.setEnabled(False)
                self.actionOpenScript.setEnabled(False)
                self.actionOpen_TestCase.setEnabled(False)
                self.actionConvertExcelToJson.setEnabled(False)
                self.actionEnable_lab.setEnabled(False)
                self.actionDisable_factory.setEnabled(False)

    def init_label_info(self):
        self.actionproduction.setText(gv.cfg.dut.test_mode)
        self.action192_168_1_101.setText(GetAllIpv4Address('10.90.'))

    def init_status_bar(self):
        with dal.database.sqlite.Sqlite(gv.DatabaseSetting) as db:
            db.execute_commit(f"SELECT VALUE  from COUNT WHERE NAME='continue_fail_count'")
            self.continue_fail_count = db.cur.fetchone()[0]
            db.execute_commit(f"SELECT VALUE  from COUNT WHERE NAME='total_pass_count'")
            self.total_pass_count = db.cur.fetchone()[0]
            db.execute_commit(f"SELECT VALUE  from COUNT WHERE NAME='total_fail_count'")
            self.total_fail_count = db.cur.fetchone()[0]
            db.execute_commit(f"SELECT VALUE  from COUNT WHERE NAME='total_abort_count'")
            self.total_abort_count = db.cur.fetchone()[0]

        self.lb_continuous_fail = QLabel(f'continuous_fail: {self.continue_fail_count}')
        self.lb_count_pass = QLabel(f'PASS: {self.total_pass_count}')
        self.lb_count_fail = QLabel(f'FAIL: {self.total_fail_count}')
        self.lb_count_abort = QLabel(f'ABORT: {self.total_abort_count}')
        try:
            self.lb_count_yield = QLabel('Yield: {:.2%}'.format(self.total_pass_count / (
                    self.total_pass_count + self.total_fail_count + self.total_abort_count)))
        except ZeroDivisionError:
            self.lb_count_yield = QLabel('Yield: 0.00%')
        self.connect_status_image = QLabel('')
        self.connect_status_txt = QLabel('')
        self.statusbar.addPermanentWidget(self.lb_count_pass, 2)
        self.statusbar.addPermanentWidget(self.lb_count_fail, 2)
        self.statusbar.addPermanentWidget(self.lb_count_abort, 2)
        self.statusbar.addPermanentWidget(self.lb_count_yield, 2)
        self.init_progress_bar()
        self.statusbar.addPermanentWidget(self.progress_bar, 8)

    def init_progress_bar(self):
        style = '''
            QProgressBar {
                border: 2px solid #000;
                border-radius: 5px;
                text-align:center;
                height: 20px;
                width:200px;
            }
            QProgressBar::chunk {
                background: #09c;
                width:1px;
            }
        '''
        self.progress_bar = QtWidgets.QProgressBar(self.statusbar)
        self.progress_bar.setRange(0, self.testcase.stepCount)
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet(style)
        self.progress_bar.setFormat('%v/%m')

    def init_signals_connect(self):
        """connect signals to slots"""
        self.mySignals.timingSignal[bool].connect(self.timing)
        self.mySignals.updateLabel[QLabel, str, int, QBrush].connect(update_label)
        self.mySignals.updateLabel[QLabel, str, int].connect(update_label)
        self.mySignals.updateLabel[QLabel, str].connect(update_label)
        self.mySignals.update_tableWidget[list].connect(self.on_updateTableWidgetRet)
        self.mySignals.textEditClearSignal[str].connect(self.on_textEditClear)
        self.mySignals.lineEditEnableSignal[bool].connect(self.lineEditEnable)
        self.mySignals.setIconSignal[QAction, QIcon].connect(on_setIcon)
        self.mySignals.updateActionSignal[QAction, QIcon].connect(updateAction)
        self.mySignals.updateActionSignal[QAction, QIcon, str].connect(updateAction)
        self.mySignals.updateStatusBarSignal[str].connect(self.updateStatusBar)
        self.mySignals.saveTextEditSignal[str].connect(self.on_actionSaveLog)
        self.mySignals.controlEnableSignal[QAction, bool].connect(controlEnable)
        self.mySignals.treeWidgetColor[QBrush, int, int, bool].connect(self.update_treeWidget_color)
        self.mySignals.threadStopSignal[str].connect(self.on_actionStop)
        self.mySignals.updateConnectStatusSignal[bool, str].connect(self.on_connect_status)
        self.mySignals.showMessageBox[str, str, int].connect(self.showMessageBox)
        self.mySignals.updateProgressBar[int].connect(self.update_progress_bar)
        self.mySignals.updateProgressBar[int, int].connect(self.update_progress_bar)
        self.mySignals.play_audio[str].connect(self.on_play_audio)

        self.actionCheckAll.triggered.connect(self.on_actionCheckAll)
        self.actionUncheckAll.triggered.connect(self.on_actionUncheckAll)
        self.actionStepping.triggered.connect(self.on_actionStepping)
        self.actionLooping.triggered.connect(self.on_actionLooping)
        self.actionExpandAll.triggered.connect(self.on_actionExpandAll)
        self.actionCollapseAll.triggered.connect(self.on_actionCollapseAll)
        self.actionBreakpoint.triggered.connect(self.on_actionBreakpoint)
        self.actionCopy.triggered.connect(self.on_actionCopy)
        self.actionPaste.triggered.connect(self.on_actionPaste)
        self.actionDelete.triggered.connect(self.on_actionDelete)

        self.actionNewSeq.triggered.connect(self.on_actionNewSequence)
        self.actionOpen_TestCase.triggered.connect(self.on_actionOpen_TestCase)
        self.actionConvertExcelToJson.triggered.connect(self.on_actionConvertExcelToJson)
        self.actionOpenScript.triggered.connect(self.on_actionOpenScript)
        self.actionSaveToScript.triggered.connect(self.on_actionSaveToScript)
        self.actionReloadScript.triggered.connect(self.on_reloadSeqs)
        self.actionConfig.triggered.connect(self.on_actionConfig)
        self.actionPrivileges.triggered.connect(self.on_actionPrivileges)
        self.actionStart.triggered.connect(self.on_actionStart)
        self.actionStop.triggered.connect(self.on_actionStop)
        self.actionOpenLog.triggered.connect(self.on_actionOpenLog)
        self.actionClearLog.triggered.connect(self.on_actionClearLog)
        self.actionLogFolder.triggered.connect(on_actionLogFolder)
        self.actionSaveLog.triggered.connect(self.on_actionSaveLog)
        self.actionCSVLog.triggered.connect(self.on_actionCSVLog)
        self.actionException.triggered.connect(on_actionException)
        self.actionEnable_lab.triggered.connect(self.on_actionEnable_lab)
        self.actionDisable_factory.triggered.connect(self.on_actionDisable_factory)
        self.actionAbout.triggered.connect(self.on_actionAbout)
        self.actionRestart.triggered.connect(self.on_actionRestart)
        self.actionPeakLin.triggered.connect(self.on_peak_lin)
        self.actionUpdates.triggered.connect(self.on_actionUpdates)

        self.lineEdit.textEdited.connect(self.on_textEdited)
        self.lineEdit.returnPressed.connect(self.on_returnPressed)
        self.treeView.customContextMenuRequested.connect(self.on_treeWidgetMenu)
        self.treeView.pressed.connect(self.on_itemPressed)
        self.pBt_start.toggled.connect(self.on_pBtToggled)
        self.tabWidget.tabBarClicked.connect(self.on_tabBarClicked)

    def init_treeView_color(self):
        for row in range(self.treeViewModel.rowCount()):
            for column in range(self.treeViewModel.columnCount()):
                suiteItem = self.treeViewModel.item(row, column)
                suiteItem.setBackground(Qt.white)
                for i in range(suiteItem.rowCount()):
                    self.treeViewModel.item(row, column).child(i).setBackground(Qt.white)

    def on_itemPressed(self, index):
        item = self.treeViewModel.itemFromIndex(index)
        if item.parent() is None:
            # self.logger.critical('itemActivate')
            self.SuiteNo = item.index().row()
            self.StepNo = -1
            self.treeView.setExpanded(self.treeViewModel.item(self.SuiteNo, 0).index(), True)
            self.actionStepping.setEnabled(False)
            self.actionEditStep.setEnabled(False)
            pp = item.text().split(' ', 1)[1]
            anchor = f'testSuite:{pp}'
            self.textEdit.scrollToAnchor(anchor)
        else:
            # self.logger.critical('itemActivate')
            self.SuiteNo = item.parent().index().row()
            self.StepNo = item.index().row()
            self.actionStepping.setEnabled(True)
            self.actionEditStep.setEnabled(True)
            # print(self.SuiteNo, self.StepNo)
            pp = item.parent().text().split(' ', 1)[1]
            cc = item.text().split(' ', 1)[1]
            anchor = f'testStep:{pp}-{cc}'
            self.textEdit.scrollToAnchor(anchor)
            if not gv.IsHide:
                self.on_actionShowStepInfo()

    def on_updateTableWidgetRet(self, result_list):
        if gv.IsHide:
            return
        if len(result_list) == 0:
            self.tableViewRetModel.clear()
            self.tableViewRetModel.setHorizontalHeaderLabels(self.tableWidgetHeader)
        else:
            itemList = []
            for i, result in enumerate(result_list):
                item = QStandardItem(str(result))
                item.setData(str(result), Qt.ToolTipRole)
                if result_list[len(result_list) - 1].lower() == 'fail':
                    item.setForeground(Qt.red)
                itemList.append(item)
            self.tableViewRetModel.appendRow(itemList)

    def on_reloadSeqs(self):
        if self.startFlag:
            QMessageBox.information(self, 'Infor', 'Test is running, can not reload!', QMessageBox.Yes)
            return
        self.logger.debug('start reload script,please wait a moment...')
        QApplication.processEvents()

        def thread_convert_and_load_script():
            if os.path.exists(self.testcase.test_script_json):
                os.chmod(self.testcase.test_script_json, stat.S_IWRITE)
                os.remove(self.testcase.test_script_json)
            self.testcase = models.testcase.TestCase(gv.ExcelFilePath, gv.cfg.station.station_name, self.logger, self,
                                                     False)
            self.testSequences = self.testcase.cloneSuites

        thread = Thread(target=thread_convert_and_load_script, daemon=True)
        thread.start()
        thread.join()
        if self.testSequences is not None:
            self.ShowTreeView(self.testSequences)
        self.logger.debug('reload finish!')
        self.SaveScriptDisableFlag = False

    def on_treeItemChanged(self, item: QStandardItem):
        if item is None:
            return
        if item.isCheckable():
            state = item.checkState()
            if item.isTristate():
                if state != Qt.PartiallyChecked:
                    self.treeItem_checkAllChild(item, True if state == Qt.Checked else False)
            else:
                self.treeItem_CheckChildChanged(item, True if state == Qt.Checked else False)

    def treeItem_checkAllChild(self, item: QStandardItem, check: bool):
        if item is None:
            return
        pNo = item.index().row()
        self.testcase.cloneSuites[pNo].isTest = check
        for i in range(0, item.rowCount()):
            self.treeItem_checkAllChild_recursion(item.child(i), check)
            self.testcase.cloneSuites[pNo].steps[i].isTest = check
        if item.isCheckable():
            item.setCheckState(Qt.Checked if check else Qt.Unchecked)

    def treeItem_checkAllChild_recursion(self, item: QStandardItem, check: bool):
        if item is None:
            return
        for i in range(0, item.rowCount()):
            self.treeItem_checkAllChild_recursion(item.child(i), check)
        if item.isCheckable():
            item.setCheckState(Qt.Checked if check else Qt.Unchecked)

    def treeItem_CheckChildChanged(self, item: QStandardItem, check: bool):
        if item is None:
            return
        siblingState = self.checkSibling(item)
        parentItem = item.parent()
        if parentItem is None:
            return
        pNo = parentItem.index().row()
        cNO = item.index().row()
        self.testcase.cloneSuites[pNo].steps[cNO].isTest = check
        if Qt.PartiallyChecked == siblingState:
            if parentItem.isCheckable() and parentItem.isTristate():
                parentItem.setCheckState(Qt.PartiallyChecked)
                self.testcase.cloneSuites[pNo].isTest = True
        elif Qt.Checked == siblingState:
            if parentItem.isCheckable():
                parentItem.setCheckState(Qt.Checked)
                self.testcase.cloneSuites[pNo].isTest = True
        else:
            if parentItem.isCheckable():
                parentItem.setCheckState(Qt.Unchecked)
                self.testcase.cloneSuites[pNo].isTest = False
        self.treeItem_CheckChildChanged(parentItem, check)

    @staticmethod
    def checkSibling(item: QStandardItem):
        parent = item.parent()
        if parent is None:
            return item.checkState()

        unCheckedCount = 0
        checkedCount = 0
        for i in range(parent.rowCount()):
            siblingItem = parent.child(i)
            state = siblingItem.checkState()
            if Qt.PartiallyChecked == state:
                return Qt.PartiallyChecked
            elif Qt.Unchecked == state:
                unCheckedCount += 1
            else:
                checkedCount += 1
            if checkedCount > 0 and unCheckedCount > 0:
                return Qt.PartiallyChecked
        if unCheckedCount > 0:
            return Qt.Unchecked
        return Qt.Checked

    def on_treeWidgetMenu(self):
        if gv.IsDebug:
            menu = QMenu(self.treeView)
            sub_menu = QMenu("Edit", menu)
            sub_menu.setObjectName("Edit")
            sub_menu.setIcon(QIcon(':/images/edit.png'))
            if not gv.IsHide:
                menu.addMenu(sub_menu)
                sub_menu.addAction(self.actionCopy)
                sub_menu.addAction(self.actionPaste)
                sub_menu.addAction(self.actionDelete)
            menu.addAction(self.actionStepping)
            menu.addAction(self.actionLooping)
            menu.addAction(self.actionBreakpoint)
            menu.addAction(self.actionCheckAll)
            menu.addAction(self.actionUncheckAll)
            menu.addAction(self.actionExpandAll)
            menu.addAction(self.actionCollapseAll)
            if not getattr(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo], 'breakpoint'):
                self.actionBreakpoint.setIcon(QIcon(':/images/StepBreakpoint.ico'))
            else:
                self.actionBreakpoint.setIcon(QIcon(':/images/BreakpointDisabled.ico'))
                self.actionBreakpoint.setText('Breakpoint Clear')
            if self.startFlag:
                self.actionStepping.setEnabled(False)
                self.actionLooping.setEnabled(False)
                self.actionCheckAll.setEnabled(False)
                self.actionUncheckAll.setEnabled(False)
                sub_menu.setEnabled(False)
            else:
                self.actionStepping.setEnabled(True)
                self.actionLooping.setEnabled(True)
                self.actionCheckAll.setEnabled(True)
                self.actionUncheckAll.setEnabled(True)
                sub_menu.setEnabled(True)
            menu.exec_(QCursor.pos())

    def on_actionCheckAll(self):
        self.ShowTreeView(self.testSequences, True)

    def on_actionUncheckAll(self):
        self.ShowTreeView(self.testSequences, False)

    def on_actionStepping(self):
        self.on_returnPressed('stepping')

    def on_actionLooping(self):
        self.FailNumOfCycleTest = 0
        self.PassNumOfCycleTest = 0
        self.IsCycle = True
        self.on_returnPressed()

    def on_actionBreakpoint(self):
        if not getattr(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo], 'breakpoint'):
            setattr(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo], 'breakpoint', True)
            self.actionBreakpoint.setIcon(QIcon(':/images/BreakpointDisabled.ico'))
            self.treeViewModel.itemFromIndex(self.treeView.currentIndex()).setIcon(QIcon(':/images/StepBreakpoint.ico'))
        else:
            setattr(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo], 'breakpoint', False)
            self.actionBreakpoint.setIcon(QIcon(':/images/StepBreakpoint.ico'))
            self.treeViewModel.itemFromIndex(self.treeView.currentIndex()).setIcon(
                QIcon(':/images/Document-txt-icon.png'))

    def on_actionOpen_TestCase(self):
        def thread_actionOpen_TestCase():
            os.startfile(self.testcase.testcasePath)

        thread = Thread(target=thread_actionOpen_TestCase, daemon=True)
        thread.start()

    def on_actionConvertExcelToJson(self):
        if self.startFlag:
            QMessageBox.information(self, 'Infor', 'Test is running, can not do it!', QMessageBox.Yes)
            return
        thread = Thread(
            target=models.loadseq.excel_convert_to_json, args=(self.testcase.testcasePath,
                                                               gv.cfg.station.station_all, self.logger), daemon=True)
        thread.start()

    def on_actionOpenScript(self):

        def actionOpenScript():
            os.startfile(self.testcase.test_script_json)

        if getattr(sys, 'frozen', False):
            return
        thread = Thread(target=actionOpenScript, daemon=True)
        thread.start()

    def on_selectStation(self):
        def select_station():
            action = self.sender()
            if isinstance(action, QAction):
                gv.cfg.station.station_name = action.text() if "(" not in action.text() else action.text()[
                                                                                             :action.text().index('(')]
                self.dut_model = ''
                self.actionunknow.setText('')
            if isinstance(action, QMenu):
                gv.cfg.station.station_name = action.title() if "(" not in action.title() else action.title()[
                                                                                               :action.title().index(
                                                                                                   '(')]
            if gv.cfg.station.station_name.startswith('ReadVer'):
                gv.IsDebug = True
            gv.cfg.station.station_no = gv.cfg.station.station_name
            self.testcase = models.testcase.TestCase(gv.ExcelFilePath, gv.cfg.station.station_name, self.logger,
                                                     self, False)
            self.mySignals.updateProgressBar[int, int].emit(self.testcase.stepFinishNum, self.testcase.sumStep)
            self.testSequences = self.testcase.cloneSuites
            self.logger.debug(f'select {self.testcase.test_script_json} finish!')

        thread = Thread(target=select_station, daemon=True)
        thread.start()
        thread.join()
        if self.testSequences is not None:
            self.ShowTreeView(self.testSequences)

    def on_selectDutModel(self):
        action = self.sender()
        if isinstance(action, QAction):
            self.dut_model = action.text() if "(" not in action.text() else action.text()[:action.text().index('(')]
            self.actionunknow.setText(self.dut_model)
        self.treeViewModel.setHorizontalHeaderLabels([f'{gv.cfg.station.station_no}_{self.dut_model}'])
        if self.lineEdit.receivers(self.lineEdit.textEdited) == 1:
            self.lineEdit.textEdited.disconnect(self.on_textEdited)

    def on_actionConfig(self):
        settings_wind = SettingsDialog(self)
        settings_wind.exec_()
        if settings_wind.isChange:
            ask = QMessageBox.question(self, "Save configuration to file?",
                                       "The configuration has been changed.Do you want to save it permanently?",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if ask == QMessageBox.Yes:
                conf.config.save_config(gv.cfg, gv.ConfigYamlPath)
        settings_wind.destroy()

    def on_peak_lin(self):
        if gv.PLin is None:
            gv.PLin = PeakLin(self.logger, self)
            gv.PLin.exec_()
        else:
            gv.PLin.show()

    def on_actionPrivileges(self):
        if gv.IsDebug:
            QMessageBox.information(self, 'Authority', 'This is lab test privileges.', QMessageBox.Yes)
        else:
            QMessageBox.information(self, 'Authority', 'This is factory test privileges.', QMessageBox.Yes)

    def on_actionStart(self):
        if self.startFlag:
            if not self.pauseFlag:
                self.pauseFlag = True
                self.actionStart.setIcon(QIcon(':/images/Start-icon.png'))
                self.pause_event.clear()
                self.treeView.blockSignals(False)
            else:
                self.pauseFlag = False
                self.actionStart.setIcon(QIcon(':/images/Pause-icon.png'))
                self.pause_event.set()
                self.treeView.blockSignals(True)
        else:
            self.on_returnPressed()

    def on_actionStop(self):
        if self.IsCycle:
            if self.startFlag:
                self.saveTestResult()
                if self.FailNumOfCycleTest == 0:
                    self.finalTestResult = True
                    self.testThread.signal[MainForm, TestStatus].emit(self, TestStatus.PASS)
                else:
                    self.testThread.signal[MainForm, TestStatus].emit(self, TestStatus.FAIL)
                self.IsCycle = False
        else:
            self.testThread.signal[MainForm, TestStatus].emit(self, TestStatus.ABORT)

    def on_actionClearLog(self):
        if not self.startFlag:
            self.textEdit.clear()

    def on_actionOpenLog(self):
        def thread_update():
            if os.path.exists(self.txtLogPath):
                # os.startfile(self.txtLogPath)
                QDesktopServices.openUrl(QUrl(f'file:///{ensure_path_sep(self.txtLogPath)}'))
            else:
                self.logger.warning(f"no find txt log,path:{self.txtLogPath}")

        thread = Thread(target=thread_update, daemon=True)
        thread.start()

    def on_actionCSVLog(self):
        def thread_update():
            if os.path.exists(self.testcase.csvFilePath):
                os.startfile(self.testcase.csvFilePath)
            else:
                self.logger.warning(f"no find CSV log,path:{self.testcase.csvFilePath}")

        thread = Thread(target=thread_update, daemon=True)
        thread.start()

    def on_actionSaveLog(self, info):
        def thread_update():
            if info == 'rename':
                rename_log = self.txtLogPath.replace('logging',
                                                     str(self.finalTestResult).upper()).replace('details',
                                                                                                str(self.testcase.errorDetailsFirstFail))
                self.logger.debug(f"rename test log to: {rename_log}")
                self.fileHandle.close()
                os.rename(self.txtLogPath, rename_log)
                self.txtLogPath = rename_log
                myScreenshot = pyautogui.screenshot(region=(self.x(), self.y(), self.width(), self.height() + 50))
                ScreenshotPhoto = self.txtLogPath.replace('.txt', '.png')
                myScreenshot.save(ScreenshotPhoto)
            else:
                self.txtLogPath = rf'{gv.LogFolderPath}{os.sep}{str(self.finalTestResult).upper()}_{self.SN}_' \
                                  rf'{self.testcase.errorDetailsFirstFail}_{time.strftime("%H-%M-%S")}.txt'
                content = self.textEdit.toPlainText()

                with open(self.txtLogPath, 'wb') as f:
                    f.write(content.encode('utf8'))
                self.logger = LogPrint('debug', gv.CriticalLog, gv.ErrorsLog).logger
                self.logger.debug(f"Save test log path.{self.txtLogPath}")

        thread = Thread(target=thread_update, daemon=True)
        thread.start()

    def on_actionEnable_lab(self):
        gv.IsDebug = True
        self.actionPrivileges.setIcon(QIcon(':/images/UnlockedObject.ico'))
        self.debug_switch(gv.IsDebug)

    def on_actionDisable_factory(self):
        gv.IsDebug = False
        self.actionPrivileges.setIcon(QIcon(':/images/LockedObject.ico'))
        self.debug_switch(gv.IsDebug)

    def on_actionAbout(self):
        QMessageBox.about(self, 'About', 'Python3.11+PyQt5\nTechnical support: StevenShen\nWeChat:chenhlzqbx')

    def on_actionUpdates(self):
        pass

    def on_actionRestart(self):
        def thread_update():
            process_name = psutil.Process(os.getpid()).name()
            run_cmd(self.logger, rf'{gv.CurrentDir}{os.sep}tool{os.sep}restart -n {process_name} -p {process_name}')

        thread = Thread(target=thread_update)
        ask = QMessageBox.question(self, "Restart Application?",
                                   "Do you want restart this program?",
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if ask == QMessageBox.Yes:
            thread.start()

    def debug_switch(self, isDebug: bool):
        self.actionConvertExcelToJson.setEnabled(isDebug)
        self.actionSaveToScript.setEnabled(isDebug)
        self.actionReloadScript.setEnabled(isDebug)
        self.actionStart.setEnabled(isDebug)
        self.actionStop.setEnabled(isDebug)
        self.actionClearLog.setEnabled(isDebug)
        self.actionSaveLog.setEnabled(isDebug)
        self.actionConfig.setEnabled(isDebug)
        self.ShowTreeView(self.testSequences)

    def eventFilter(self, obj: 'QObject', event: 'QEvent') -> bool:
        if obj == self.tableViewStepProp and event.type() == QEvent.Type.ContextMenu:
            self.stepMenu.exec(QCursor.pos())
        # if obj == self.tableViewRet and event.type() == QEvent.Type.ContextMenu:
        #     rect = QRect(QCursor.pos().x(), QCursor.pos().y(), 50, 10)
        #     QToolInfo = self.tableViewRetModel.data(self.tableViewRetModel.index(self.tableViewRet.currentIndex().row(),
        #                                                                          self.tableViewRet.currentIndex().column()))
        #     QToolTip.showText(QCursor.pos(), QToolInfo, self, rect, 3000)
        return super().eventFilter(obj, event)

    def on_actionShowStepInfo(self):
        if self.tabWidget.currentWidget() != self.stepInfo:
            self.tabWidget.setCurrentWidget(self.stepInfo)
        step_obj = self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo]
        self.model = QStandardItemModel(0, 2, self.tableViewStepProp)
        self.model.itemChanged.connect(self.on_stepInfoEdit)
        for prop_name in gv.StepAttr:
            prop_value = getattr(step_obj, prop_name)
            self.model.setHorizontalHeaderLabels(['Property', 'Value'])
            prop_name_item = QStandardItem(prop_name)
            prop_name_item.setBackground(QColor('#E6E6E6'))
            prop_name_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            prop_value_item = QStandardItem(str(prop_value))
            if prop_value is None:
                prop_value_item.setBackground(Qt.lightGray)
            if prop_name == 'SuiteName' and self.StepNo != 0:
                prop_value_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                prop_value_item.setBackground(Qt.lightGray)
            self.model.appendRow([prop_name_item, prop_value_item])
        self.tableViewStepProp.setModel(self.model)

    def on_stepInfoEdit(self, item: QStandardItem):
        prop_name = self.model.data(self.model.index(item.row(), 0))
        prop_value = item.text()
        step_obj = self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo]
        try:
            T = (type(getattr(step_obj, prop_name)))
            if prop_value == 'None':
                prop_value = None
                setattr(step_obj, prop_name, prop_value)
            else:
                setattr(step_obj, prop_name, T(prop_value))
        except ValueError:
            item.setBackground(Qt.red)
            raise
        else:
            item.setBackground(Qt.white)
        self.header_new = gv.StepAttr
        self.actionSaveToScript.setEnabled(True)

    def on_tabBarClicked(self, index):
        if self.tabWidget.tabText(index) == 'Variables':
            if self.TestVariables is None:
                return
            model = QStandardItemModel(0, 2, self.tableViewVar)
            for prop in self.TestVariables:
                if prop[0] != 'Config':
                    model.setHorizontalHeaderLabels(['Variable', 'Value'])
                    prop_name = QStandardItem(prop[0])
                    prop_name.setBackground(QColor('#E6E6E6'))
                    prop_name.setFlags(Qt.ItemIsSelectable)
                    prop_value = QStandardItem(str(prop[1]))
                    model.appendRow([prop_name, prop_value])
            self.tableViewVar.setModel(model)

    def on_actionSaveToScript(self):
        # if self.SaveScriptDisableFlag:
        #     QMessageBox.information(self, 'Infor', 'Please save it before start test!', QMessageBox.Yes)
        #     self.actionSaveToScript.setEnabled(False)
        #     return
        if not self.header_new:
            raise ValueError('the list of excel header_new cannot be empty!')

        def SaveToScript():
            step_value = []
            sheet_name = gv.cfg.station.station_name
            workbook = openpyxl.load_workbook(gv.ExcelFilePath)
            try:
                ws = workbook[sheet_name]
            except KeyError:
                workbook.create_sheet(sheet_name)
                ws = workbook[sheet_name]
            try:
                sheet_index = workbook.sheetnames.index(sheet_name)
                workbook.active = sheet_index
                for table in ws.tables.items():
                    try:
                        del ws.tables[table[0]]
                    except KeyError:
                        pass
                ws.delete_rows(idx=0, amount=self.testcase.stepCount * 3)
                ws.append(self.header_new)
                for suit in self.testcase.cloneSuites:
                    for step in suit.steps:
                        for item in self.header_new:
                            value = getattr(step, item)
                            if item == 'SuiteName' and suit.steps.index(step) != 0:
                                value = ''
                            step_value.append('' if value is None else value)
                        ws.append(step_value)
                        step_value = []
                tab = Table(displayName=sheet_name,
                            ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
                ws.views.sheetView[0].zoomScale = 80
                column_width_autofit(ws)
                # ws.protection.sheet = True
                # ws.protection.enable()
                # ws.protection.password = '....'
            except Exception as e:
                self.logger.error(f'{currentframe().f_code.co_name}:{e}')
                QMetaObject.invokeMethod(
                    self,
                    'showMessageBox',
                    Qt.BlockingQueuedConnection,
                    QtCore.Q_RETURN_ARG(QMessageBox.StandardButton),
                    QtCore.Q_ARG(str, 'ERROR!'),
                    QtCore.Q_ARG(str, f'{currentframe().f_code.co_name}:{e}'),
                    QtCore.Q_ARG(int, 4))
            else:
                try:
                    workbook.save(gv.ExcelFilePath)
                    self.logger.debug(f'sync save to excel:{gv.ExcelFilePath}')
                    self.on_reloadSeqs()
                except PermissionError as e:
                    self.logger.error(f'{e},Please close the excel file first.')
                except Exception as e:
                    self.logger.error(f'{e}')
                    raise
                else:
                    self.actionSaveToScript.setEnabled(False)

        thread = Thread(target=SaveToScript, daemon=True)
        thread.start()

    def on_actionExpandAll(self):
        self.treeView.expandAll()

    def on_actionCollapseAll(self):
        self.treeView.collapseAll()

    def on_connect_status(self, flag: bool, strs):
        if flag:
            self.connect_status_image.setPixmap(QPixmap(":/images/connect_ok.png"))
        else:
            self.connect_status_image.setPixmap(QPixmap(":/images/disconnect-icon.png"))
        self.connect_status_txt.setText(strs)
        self.statusbar.addPermanentWidget(self.connect_status_image, 1)
        self.statusbar.addPermanentWidget(self.connect_status_txt, 2)

    def ShowTreeView(self, sequences=None, checkall=None):
        if sequences is None:
            return
        self.treeViewModel.clear()
        dut_model = '' if self.dut_model == '' else '_' + self.dut_model
        self.treeViewModel.setHorizontalHeaderLabels([f'{gv.cfg.station.station_no}{dut_model}'])
        itemRoot = self.treeViewModel.invisibleRootItem()

        for suite in sequences:
            if gv.IsHide:
                suiteItem = QStandardItem(QIcon(':/images/folder-icon.png'), f'{suite.index + 1}. suite')
            else:
                suiteItem = QStandardItem(QIcon(':/images/folder-icon.png'), f'{suite.index + 1}. {suite.name}')

            if checkall:
                suiteItem.setCheckState(Qt.Checked)
                suite.isTest = True
            else:
                if checkall is None:
                    suiteItem.setCheckState(Qt.Checked if suite.isTest else Qt.Unchecked)
                else:
                    suiteItem.setCheckState(Qt.Unchecked)
                    suite.isTest = False
            if gv.IsDebug:
                suiteItem.setFlags(Qt.ItemIsSelectable | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            else:
                suiteItem.setFlags(Qt.ItemIsSelectable)
            suiteItem.setTristate(True)
            itemRoot.appendRow(suiteItem)

            for step in suite.steps:
                if gv.IsHide:
                    stepItem = QStandardItem(QIcon(':/images/Document-txt-icon.png'), f'{step.index + 1}) step')
                else:
                    stepItem = QStandardItem(QIcon(':/images/Document-txt-icon.png'),
                                             f'{step.index + 1}) {step.StepName}')
                if checkall:
                    stepItem.setCheckState(Qt.Checked)
                    step.isTest = True
                else:
                    if checkall is None:
                        stepItem.setCheckState(Qt.Checked if step.isTest else Qt.Unchecked)
                    else:
                        stepItem.setCheckState(Qt.Unchecked)
                        step.isTest = False
                if gv.IsDebug:
                    stepItem.setFlags(Qt.ItemIsSelectable | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                else:
                    stepItem.setFlags(Qt.ItemIsSelectable)
                self.setStepIcon(step, stepItem)
                suiteItem.appendRow(stepItem)
        self.treeView.setModel(self.treeViewModel)
        # self.treeView.expandAll()
        self.treeView.setExpanded(self.treeViewModel.item(0, 0).index(), True)
        self.treeViewModel.itemChanged.connect(self.on_treeItemChanged)  # on_itemChanged
        self.tabWidget.setCurrentWidget(self.result)

    @staticmethod
    def setStepIcon(step, stepItem):
        if step.Keyword == 'MessageBoxShow' or step.Keyword == 'QInputDialog' or step.Keyword == 'DialogInput':
            stepItem.setIcon(QIcon(':/images/MsgBox.ico'))
        else:
            if step.IfElse == 'if' or step.IfElse == '&if' or step.IfElse == '||if':
                stepItem.setIcon(QIcon(':/images/NI_If.ico'))
            elif step.IfElse == 'elif':
                stepItem.setIcon(QIcon(':/images/NI_ElseIf.ico'))
            elif step.IfElse == 'else':
                stepItem.setIcon(QIcon(':/images/NI_Else.ico'))
            else:
                if str_to_int(step.For)[0]:
                    stepItem.setIcon(QIcon(':/images/NI_For.ico'))
                elif step.For == 'do' or step.For == 'while':
                    stepItem.setIcon(QIcon(':/images/NI_DoWhile.ico'))
                elif step.For == 'whiledo':
                    stepItem.setIcon(QIcon(':/images/NI_While.ico'))
                elif not IsNullOrEmpty(step.For) and step.For.startswith('end'):
                    stepItem.setIcon(QIcon(':/images/NI_End.ico'))
                else:
                    stepItem.setIcon(QIcon(':/images/Document-txt-icon.png'))

    # @QtCore.pyqtSlot(QBrush, int, int, bool)
    def update_treeWidget_color(self, color: QBrush, suiteNO_: int, stepNo_: int = -1, allChild=False):
        if stepNo_ == -1:
            if self.IsCycle or not self.startFlag:
                return
            self.treeView.setExpanded(self.treeViewModel.item(suiteNO_, 0).index(), True)
            self.treeViewModel.item(suiteNO_, 0).setBackground(color)
            if allChild:
                for i in range(self.treeViewModel.item(suiteNO_, 0).rowCount()):
                    self.treeViewModel.item(suiteNO_, 0).child(i).setBackground(0, color)
        else:
            self.treeViewModel.item(suiteNO_, 0).child(stepNo_, 0).setBackground(color)
            self.treeView.scrollTo(self.treeViewModel.item(suiteNO_, 0).child(stepNo_, 0).index(),
                                   hint=QAbstractItemView.EnsureVisible)

    @QtCore.pyqtSlot(str, str, int, result=QMessageBox.StandardButton)
    def showMessageBox(self, title, text, level=2):
        if level == 0:
            return QMessageBox.information(self, title, text, QMessageBox.Yes)
        elif level == 1:
            return QMessageBox.warning(self, title, text, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        elif level == 2:
            aa = QMessageBox.question(self, title, text, QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            return aa
        elif level == 3:
            return QMessageBox.about(self, title, text)
        else:
            return QMessageBox.critical(self, title, text, QMessageBox.Yes)

    @QtCore.pyqtSlot(str, str, result=list)
    def showQInputDialog(self, title, label):
        results = QInputDialog.getText(self, title, label)
        return list(results)

    def lineEditEnable(self, isEnable):
        self.lineEdit.setEnabled(isEnable)
        if isEnable:
            self.lineEdit.setText('')
            self.lineEdit.setFocus()

    def on_textEditClear(self, info):
        self.logger.debug(f'{currentframe().f_code.co_name}:{info}')
        self.textEdit.clear()

    def updateStatusBar(self, info):
        self.logger.debug(f'{currentframe().f_code.co_name}:{info}')
        with dal.database.sqlite.Sqlite(gv.DatabaseSetting) as db:
            db.execute_commit(f"UPDATE COUNT SET VALUE='{self.continue_fail_count}' where NAME ='continue_fail_count'")
            db.execute_commit(f"UPDATE COUNT SET VALUE='{self.total_pass_count}' where NAME ='total_pass_count'")
            db.execute_commit(f"UPDATE COUNT SET VALUE='{self.total_fail_count}' where NAME ='total_fail_count'")
            db.execute_commit(f"UPDATE COUNT SET VALUE='{self.total_abort_count}' where NAME ='total_abort_count'")
        self.lb_continuous_fail.setText(f'continuous_fail: {self.continue_fail_count}')
        self.lb_count_pass.setText(f'PASS: {self.total_pass_count}')
        self.lb_count_fail.setText(f'FAIL: {self.total_fail_count}')
        self.lb_count_abort.setText(f'ABORT: {self.total_abort_count}')
        try:
            self.lb_count_yield.setText('Yield: {:.2%}'.format(self.total_pass_count / (
                    self.total_pass_count + self.total_fail_count + self.total_abort_count)))
        except ZeroDivisionError:
            self.lb_count_yield.setText('Yield: 0.00%')
        # QApplication.processEvents()

    def update_progress_bar(self, value: int, _range: int = None):
        if _range is not None:
            self.progress_bar.setRange(0, _range)
        self.progress_bar.setValue(value)

    def on_play_audio(self, path):
        player = QMediaPlayer(self)
        qurl = QUrl.fromLocalFile(path)
        qmusic = QMediaContent(qurl)
        player.setMedia(qmusic)
        player.setVolume(100)
        player.play()

    def timerEvent(self, a):
        self.mySignals.updateLabel[QLabel, str, int].emit(self.lb_errorCode, str(self.sec), 20)
        self.sec += 1

    def UpdateContinueFail(self, testResult: bool):
        if gv.IsDebug or gv.cfg.dut.test_mode.lower() == 'debug':
            return
        if testResult:
            self.continue_fail_count = 0
        else:
            self.continue_fail_count += 1

    def ContinuousFailReset_Click(self):
        """连续fail超过规定值需要TE确认问题并输入密码后才能继续测试"""
        text, ok = QInputDialog.getText(self, 'Reset', 'Please input Reset Password:', echo=QLineEdit.Password)
        if ok:
            if text == 'test123':
                self.continue_fail_count = 0
                self.lb_continuous_fail.setText(f'continuous_fail: {self.continue_fail_count}')
                self.lb_continuous_fail.setStyleSheet(
                    f"background-color:{self.statusbar.palette().window().color().name()};")
                return True
            else:
                QMessageBox.critical(self, 'ERROR!', 'wrong password!')
                return False
        else:
            return False

    def CheckContinueFailNum(self):
        with dal.database.sqlite.Sqlite(gv.DatabaseSetting) as db:
            db.execute_commit(f"SELECT VALUE  from COUNT WHERE NAME='continue_fail_count'")
            self.continue_fail_count = db.cur.fetchone()[0]
            self.logger.debug(str(self.continue_fail_count))
        if self.continue_fail_count >= gv.cfg.station.continue_fail_limit:
            self.lb_continuous_fail.setStyleSheet(f"background-color:red;")
            if gv.IsDebug:
                return True
            else:
                return self.ContinuousFailReset_Click()
        else:
            self.lb_continuous_fail.setStyleSheet(
                f"background-color:{self.statusbar.palette().window().color().name()};")
            return True

    def on_textEdited(self):

        def JudgeProdMode():
            """通过SN判断机种"""
            sn = self.lineEdit.text()
            if sn[0] == 'J' or sn[0] == '6':
                self.dut_model = gv.cfg.dut.dut_models[0]
            elif sn[0] == 'N' or sn[0] == '7':
                self.dut_model = gv.cfg.dut.dut_models[1]
            elif sn[0] == 'Q' or sn[0] == '8':
                self.dut_model = gv.cfg.dut.dut_models[2]
            elif sn[0] == 'S' or sn[0] == 'G':
                self.dut_model = gv.cfg.dut.dut_models[3]
            else:
                self.dut_model = 'unknown'
                # if not gv.IsDebug:
                #     raise Exception('dut_model is unknown!!')
            self.actionunknow.setText(self.dut_model)

        """验证dut sn的正则规则"""
        if JudgeProdMode() != 'unknown' and not gv.IsDebug:
            reg = QRegExp(gv.cfg.dut.dut_regex[self.dut_model])
            pValidator = QRegExpValidator(reg, self)
            self.lineEdit.setValidator(pValidator)

    def on_returnPressed(self, stepping_flag=None):
        if stepping_flag is not None:
            self.SingleStepTest = True
        else:
            self.SingleStepTest = False
        if not gv.IsDebug and (self.dut_model == 'unknown' or len(self.lineEdit.text()) != gv.cfg.dut.sn_len):
            str_info = f'无法根据SN判断机种或者SN长度不对! 扫描:{len(self.lineEdit.text())},规定:{gv.cfg.dut.sn_len}.'
            QMetaObject.invokeMethod(self, 'showMessageBox', Qt.AutoConnection,
                                     QtCore.Q_RETURN_ARG(QMessageBox.StandardButton),
                                     QtCore.Q_ARG(str, 'JudgeMode!'),
                                     QtCore.Q_ARG(str, str_info),
                                     QtCore.Q_ARG(int, 5))
            return
        # if not self.CheckContinueFailNum() and not gv.IsDebug:
        # return

        if gv.IsDebug:
            if not self.SingleStepTest:
                self.init_treeView_color()
        else:
            self.testSequences = copy.deepcopy(self.testcase.originalSuites)
            self.testcase.cloneSuites = self.testSequences
            self.ShowTreeView(self.testSequences)
        self.test_initialize(self.lineEdit.text())

    def test_initialize(self, SN):
        """测试变量初始化"""
        self.init_variable(SN)
        gv.lg = LogPrint(self.txtLogPath.replace('\\', '/'), gv.CriticalLog, gv.ErrorsLog)
        self.logger = gv.lg.logger
        gv.InitCreateDirs(self.logger)
        self.init_textEditHandler()
        self.testcase.logger = self.logger
        self.testcase.stepFinishNum = 0
        self.testcase.startTimeJson = datetime.now()
        self.testcase.jsonObj = models.product.JsonObject(SN, gv.cfg.station.station_no,
                                                          gv.cfg.dut.test_mode, gv.cfg.dut.qsdk_ver, gv.VERSION)
        self.testcase.mesPhases = models.product.MesInfo(SN, gv.cfg.station.station_no, gv.VERSION)
        self.testcase.daqDataPath = rf'{gv.OutPutPath}\{gv.cfg.station.station_no}_DAQ_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}.csv'
        self.lb_failInfo.setHidden(True)
        self.lb_testTime.setHidden(True)
        if not self.testThread.isRunning():
            self.testThread = TestThread(self)
            self.testThread.start()
        self.testThread.signal[MainForm, TestStatus].emit(self, TestStatus.START)
        self.tabWidget.setCurrentWidget(self.result)

    def saveTestResult(self):
        def thread_update():
            reportPath = fr'{gv.OutPutPath}\result.csv'
            create_csv_file(self.logger, reportPath, self.tableWidgetHeader)
            if os.path.exists(reportPath):
                all_rows = []
                for row in range(self.tableViewRetModel.rowCount()):
                    row_data = []
                    for column in range(self.tableViewRetModel.columnCount()):
                        item = self.tableViewRetModel.item(row, column)
                        if item is not None:
                            row_data.append(item.text())
                    all_rows.append(row_data)

                with open(reportPath, 'a', newline='') as stream:
                    writer = csv.writer(stream)
                    writer.writerows(all_rows)
            if self.logger is not None:
                self.logger.debug(f'saveTestResult to:{reportPath}')

        thread = Thread(target=thread_update, daemon=True)
        thread.start()
        thread.join()

    def open_camera(self):
        if not gv.cfg.station.auto_scan:
            return
        if gv.cfg.station.station_name == "CCT":
            return
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            self.logger.debug("Cannot open camera or no video devices!")
            return
        else:
            self.logger.debug('open camera ok.')
        video = threading.Thread(target=self.auto_scan, daemon=True)  # 將 OpenCV 的部分放入 threading 裡執行
        video.start()

    def auto_scan(self):
        try:
            fixComRcv = ''
            # # self.bt_photo.setEnabled(True)
            while self.autoScanFlag:
                if self.startFlag:
                    continue
                # fixComRcv += self.testcase.FixSerialPort.read()
                # # self.tabWidget.setCurrentWidget(self.video)
                ret, frame = self.cap.read()
                if not ret:
                    print("Cannot receive frame")
                    break
                frame = cv2.resize(frame, (self.lb_video.height(), self.lb_video.width()))
                cv2.imwrite('autoScan.jpg', frame)
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # RGB
                height, width, channel = frame.shape
                img = QImage(frame, width, height, channel * width, QImage.Format_RGB888)
                # # self.lb_video.setPixmap(QPixmap.fromImage(img))
                reader = zxing.BarCodeReader()
                result = reader.decode("autoScan.jpg").parsed
                # self.bt_photo.setEnabled(False)
                # if 'AT+SCAN' in fixComRcv:
                #     pass
                # else:
                #     if result is None or result == self._lastSn:
                #         continue
                self.lineEdit.setText(result)
                # if not self.lineEdit.Text() == result:
                #     continue
                # # self.tabWidget.setCurrentWidget(self.result)
                self.lineEdit.returnPressed.emit()
                self._lastSn = result
                fixComRcv = ''
                time.sleep(0.1)

        except RuntimeError:
            print('RuntimeError')
        except Exception as e:
            print(e)

    @staticmethod
    def sinWave(i=0):
        fig = plt.figure(figsize=(6, 4), dpi=100)
        ax = plt.axes(xlim=(0, 2), ylim=(-2, 2))
        line, = ax.plot([], [])
        line.set_data([], [])
        x = np.linspace(0, 2, 100)
        y = np.sin(5 * np.pi * (x - 0.01 * i))
        line.set_data(x, y)
        plt.close()
        return fig

    def count(self):
        self.t = self.t + 5
        self.canvas = FigureCanvas(self.sinWave(self.t))
        self.graphic_scene.clear()
        self.graphic_scene.addWidget(self.canvas)

    def on_pBtToggled(self, check: bool):
        if check:
            self.QtimerID = QTimer()
            self.QtimerID.timeout.connect(self.count)
            self.QtimerID.start(50)
            self.pBt_start.setText('Stop')
        else:
            self.QtimerID.stop()
            self.pBt_start.setText('Start')

    def on_actionDelete(self):
        if self.StepNo == -1:
            del self.testcase.cloneSuites[self.SuiteNo]
        else:
            if len(self.testcase.cloneSuites[self.SuiteNo].steps) == 1:
                del self.testcase.cloneSuites[self.SuiteNo]
                self.SuiteNo = 0
            else:
                del_step = copy.deepcopy(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo])
                del self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo]
                if self.StepNo == 0:
                    self.testcase.cloneSuites[self.SuiteNo].steps[0].SuiteName = del_step.SuiteName
                    self.testcase.cloneSuites[self.SuiteNo].steps[0].index = 0
        self.ShowTreeView(self.testSequences)
        self.actionSaveToScript.setEnabled(True)
        self.treeView.setExpanded(self.treeViewModel.item(self.SuiteNo, 0).index(), True)
        self.treeView.scrollTo(self.treeViewModel.item(self.SuiteNo, 0).index(), hint=QAbstractItemView.EnsureVisible)

    def on_actionCopy(self):
        if self.StepNo == -1:
            self.suitClipboard = copy.deepcopy(self.testcase.cloneSuites[self.SuiteNo])
        else:
            self.stepClipboard = copy.deepcopy(self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo])
            if self.StepNo == 0:
                self.stepClipboard.SuiteName = ''
        self.actionPaste.setEnabled(True)

    def on_actionPaste(self):
        if self.StepNo == -1:
            if self.suitClipboard is not None:
                self.testcase.cloneSuites.insert(self.SuiteNo, self.suitClipboard)
                self.suitClipboard = None
            elif self.stepClipboard is not None:
                self.testcase.cloneSuites[self.SuiteNo].steps.insert(0, self.stepClipboard)
                self.testcase.cloneSuites[self.SuiteNo].steps[0].SuiteName = \
                    self.testcase.cloneSuites[self.SuiteNo].steps[1].SuiteName
                self.testcase.cloneSuites[self.SuiteNo].steps[0].index = 0
                self.testcase.cloneSuites[self.SuiteNo].steps[1].index = 1
                self.testcase.cloneSuites[self.SuiteNo].steps[1].SuiteName = ''
                self.stepClipboard = None
        else:
            if self.stepClipboard is not None:
                self.testcase.cloneSuites[self.SuiteNo].steps.insert(self.StepNo + 1, self.stepClipboard)
                self.stepClipboard = None
        self.ShowTreeView(self.testSequences)
        self.actionPaste.setEnabled(False)
        self.actionSaveToScript.setEnabled(True)
        self.treeView.setExpanded(self.treeViewModel.item(self.SuiteNo, 0).index(), True)
        self.treeView.scrollTo(self.treeViewModel.item(self.SuiteNo, 0).index(), hint=QAbstractItemView.EnsureVisible)

    def on_actionNewSequence(self):
        station_name, ok = QInputDialog.getText(self, 'New TestSequences', 'test station name:')
        if ok:
            test_script_json = rf'{gv.ScriptFolder}\{station_name}.json'
            if os.path.exists(test_script_json):
                QMessageBox.critical(None, 'ERROR!', '{station_name} have existed!!', QMessageBox.Yes)
                return
            else:
                gv.cfg.station.station_name = station_name
                gv.cfg.station.station_no = gv.cfg.station.station_name
                os.system(f'copy {gv.ScriptFolder}\\sample.json {gv.ScriptFolder}\\{gv.cfg.station.station_name}.json')
            self.testcase = models.testcase.TestCase(gv.ExcelFilePath, gv.cfg.station.station_name, self.logger, self,
                                                     False, False)
            self.testSequences = self.testcase.cloneSuites
            gv.cfg.station.station_all.append(station_name)
            conf.config.save_config(gv.cfg, gv.ConfigYamlPath)
            self.ShowTreeView(self.testSequences)
            self.logger.debug(f'new {station_name} test Sequences finish!')

    def on_stepDeleteRow(self):
        prop_name = self.model.data(self.model.index(self.tableViewStepProp.currentIndex().row(), 0))
        step_obj = self.testcase.cloneSuites[self.SuiteNo].steps[self.StepNo]
        try:
            setattr(step_obj, prop_name, None)
            self.model.removeRow(self.tableViewStepProp.currentIndex().row())
            gv.StepAttr.remove(prop_name)
            self.header_new = gv.StepAttr
            self.actionSaveToScript.setEnabled(True)
        except:
            raise

    def on_stepInsertRow(self):
        pass
        # pname = 'propName'
        # pvalue = 'value'
        # row = self.tableViewStepProp.currentIndex().row()
        # self.model.insertRow(row, [QStandardItem(pname), QStandardItem(pvalue)])
        # gv.StepAttr.insert(row, pname)
        # step_obj = self.testcase.clone_suites[self.SuiteNo].steps[self.StepNo]
        # setattr(step_obj, pname, pvalue)
        # self.header_new = gv.StepAttr
        # self.actionSaveToScript.setEnabled(True)


if __name__ == "__main__":
    pass
    # app = QApplication([])
    # mainWin = MainForm()
    # mainWin.show()
    # app.exec_()
