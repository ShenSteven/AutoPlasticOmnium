#!/usr/bin/env python
# coding: utf-8
"""
@File   : load seq.py
@Author : Steven.Shen
@Date   : 2021/10/15
@Desc   : 
"""
import inspect
import os
import re
import sqlite3
import stat
import json
import sys
from inspect import currentframe
from types import NoneType
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
import models.suite
import models.step
import dataaccess.sqlite
from common.basicfunc import IsNullOrEmpty, get_sha256, get_line_list
import conf.globalvar as gv


def save_keywords_to_txt(path, wpath):
    with open(path, 'r', encoding='utf-8') as rf:
        readall = rf.read()
        SubStr1 = "register\\('"
        SubStr2 = "'\\)"
        keywords = re.findall(f'{SubStr1}(.*?){SubStr2}', readall)
        print(keywords)
        if os.path.exists(wpath):
            os.remove(wpath)
        for item in keywords:
            with open(wpath, 'a') as wf:
                wf.write(f'{item}\n')
        return keywords


def excel_convert_to_json(testcase_path_excel, all_stations, logger):
    logger.debug("Start convert excel testcase to json script,please wait a moment...")
    if not getattr(sys, 'frozen', False):
        gv.Keywords = save_keywords_to_txt(rf'{gv.CurrentDir}{os.sep}models{os.sep}keyword.py',
                                           rf'{gv.CurrentDir}{os.sep}conf{os.sep}keywords.txt')
    else:
        gv.Keywords = get_line_list(rf'{gv.CurrentDir}{os.sep}conf{os.sep}keywords.txt')
    for station in all_stations:
        load_testcase_from_excel(testcase_path_excel, station, rf"{gv.ScriptFolder}{os.sep}{station}.json", logger)
    logger.debug("convert finish!")


def load_testcase_from_excel(testcase_path, sheet_name, json_path, logger):
    """load test sequence form a sheet in Excel and return the suites sequences list,
    if success,serialize the suites list to json.
    :param logger: logger handle
    :param testcase_path: the path of Excel
    :param sheet_name: the sheet test_name of Excel
    :param json_path:  serialize and save to json path
    :return : temp_suite[] list
    """
    workbook = None
    temp_suite = None
    suites_list = []
    headers = []
    temp_suite_name = ""
    step_count = 0
    try:
        print(f'start load_testcase_from_excel sheet:{sheet_name}...')
        workbook = load_workbook(testcase_path, read_only=True)
        worksheet = workbook[sheet_name]
        for i in list(worksheet.rows)[0]:  # 获取表头，第一行
            headers.append(i.value)

        for i in range(1, worksheet.max_row):  # 一行行的读取excel
            line = list(worksheet.rows)[i]
            if IsNullOrEmpty(line[1].value):  # StepName为空停止解析
                break
            if not IsNullOrEmpty(line[0].value):  # 实例化test suite
                temp_suite_name = line[0].value
                temp_suite = models.suite.TestSuite(line[0].value, len(suites_list))
                suites_list.append(temp_suite)
            # 给step对象属性赋值
            test_step = models.step.Step()
            step_count += 1
            for header, cell in dict(zip(headers, line)).items():
                test_step.index = temp_suite.totalNum
                # test_step.suiteIndex = temp_suite.index
                if not hasattr(test_step, header):
                    setattr(test_step, header, '')
                setattr(test_step, header, '' if type(cell.value) is NoneType else str(cell.value).strip())
                if temp_suite.totalNum == 0:
                    test_step.SuiteName = temp_suite_name
            # else:
            #     setattr(test_step, 'NeverUsed', None)

            temp_suite.totalNum += 1
            temp_suite.steps.append(test_step)
    except Exception as e:
        QMessageBox.critical(None, 'ERROR!', f'{currentframe().f_code.co_name}:{e} ', QMessageBox.Yes)
        sys.exit(e)
    else:
        serialize_to_json(suites_list, json_path, logger)
    finally:
        workbook.close()


def param_wrapper_verify_sha256(flag):
    def wrapper_verify_sha2562(fun):
        """get sha256 of json script in oder to verify MD5"""
        sig = inspect.signature(fun)

        def inner(*args, **kwargs):
            sha256 = ''
            bound_args = sig.bind(*args, **kwargs)
            bound_args.apply_defaults()
            if flag and bound_args.args[1]:
                with dataaccess.sqlite.Sqlite(gv.DatabaseSetting) as db:
                    file_name = os.path.basename(bound_args.args[0])
                    db.execute_commit(f"SELECT SHA256  from SHA256_ENCRYPTION WHERE NAME='{file_name}'")
                    result = db.cur.fetchone()
                    if result:
                        sha256 = result[0]
                        # print(f"{file_name}   dbSHA:{sha256}")
                JsonSHA = get_sha256(bound_args.args[0])
                # print(f"{file_name} jsonSHA:{JsonSHA}")
                if sha256 == JsonSHA:
                    result = fun(*bound_args.args, **bound_args.kwargs)
                else:
                    error_info = f'{currentframe().f_code.co_name}:script {bound_args.args[0]} has been tampered!'
                    QMessageBox.critical(None, 'ERROR!', error_info, QMessageBox.Yes)
                    sys.exit(0)
            else:
                result = fun(*bound_args.args, **bound_args.kwargs)
            return result

        return inner

    return wrapper_verify_sha2562


@param_wrapper_verify_sha256(True)
def load_testcase_from_json(json_path, isVerify=True):
    try:
        """Deserialize form json.
        :param json_path: json file path.
        :return:object
        """
        step_count = 0
        headers = []
        with open(json_path, 'r') as rf:
            sequences_dict = json.load(rf)
        sequences_obj_list = []
        for suit_dict in sequences_dict:
            step_obj_list = []
            for step_dict in suit_dict['steps']:
                step_obj = models.step.Step(step_dict)
                step_count += 1
                step_obj_list.append(step_obj)
                if not headers:
                    param = dict(filter(lambda x: x[0][0:1].isupper() or x[0][1:2].isupper(), step_dict.items()))
                    param2 = list(filter(lambda x: param[x] is not None, param))
                    headers = list(map(lambda x: x[1:] if x.startswith('_') else x, param2))

                    items = filter(lambda x: x[0:1].isupper() or x[1:2].isupper(), step_obj.__dict__)
                    gv.StepAttr = list(map(lambda x: x[1:] if x.startswith('_') else x, items))
            suit_obj = models.suite.TestSuite(dict_=suit_dict)
            suit_obj.steps = step_obj_list
            sequences_obj_list.append(suit_obj)
        return sequences_obj_list, headers, step_count
    except Exception as e:
        QMessageBox.critical(None, 'Exception!', f'{currentframe().f_code.co_name}:{e}', QMessageBox.Yes)
        sys.exit(e)


def load_testcase_from_py(name, package='scripts'):
    try:
        """Deserialize form json.
        :param json_path: json file path.
        :return:object
        """
        step_count = 0
        headers = []
        import importlib
        b = importlib.import_module(name='.' + name, package=package)
        sequences_dict = json.loads(b.get_script_str())
        sequences_obj_list = []
        for suit_dict in sequences_dict:
            step_obj_list = []
            for step_dict in suit_dict['steps']:
                step_obj = models.step.Step(step_dict)
                step_count += 1
                step_obj_list.append(step_obj)
                if not headers:
                    param = dict(filter(lambda x: x[0][0:1].isupper() or x[0][1:2].isupper(), step_dict.items()))
                    param2 = list(filter(lambda x: param[x] is not None, param))
                    headers = list(map(lambda x: x[1:] if x.startswith('_') else x, param2))

                    items = filter(lambda x: x[0:1].isupper() or x[1:2].isupper(), step_obj.__dict__)
                    gv.StepAttr = list(map(lambda x: x[1:] if x.startswith('_') else x, items))
            suit_obj = models.suite.TestSuite(dict_=suit_dict)
            suit_obj.steps = step_obj_list
            sequences_obj_list.append(suit_obj)
        return sequences_obj_list, headers, step_count
    except Exception as e:
        QMessageBox.critical(None, 'Exception!', f'{currentframe().f_code.co_name}:{e}', QMessageBox.Yes)
        sys.exit(e)


def wrapper_save_sha256(fun):
    """save sha256 of json script in oder to verify MD5"""
    sig = inspect.signature(fun)

    def inner(*args, **kwargs):
        bound_args = sig.bind(*args, **kwargs)
        bound_args.apply_defaults()
        result = fun(*bound_args.args, **bound_args.kwargs)
        sha256 = get_sha256(bound_args.args[1])
        with dataaccess.sqlite.Sqlite(gv.DatabaseSetting) as db:
            file_name = os.path.basename(bound_args.args[1])
            try:
                db.execute_commit(f"INSERT INTO SHA256_ENCRYPTION (NAME,SHA256) VALUES ('{file_name}', '{sha256}')")
            except sqlite3.IntegrityError:
                db.execute_commit(f"UPDATE SHA256_ENCRYPTION SET SHA256='{sha256}' WHERE NAME ='{file_name}'")
        os.chmod(args[1], stat.S_IREAD)
        return result

    return inner


def default_f(o):
    try:
        _dict = o.__dict__
        _dict['myWind'] = None
        _dict['logger'] = None
        _dict['start_time'] = None
        _dict['finish_time'] = None
        if 'steps' not in _dict:
            if _dict['_index'] != 0:
                _dict['_SuiteName'] = ''
            _dict['start_time_json'] = None
            _dict['testValue'] = None
        return _dict
    except AttributeError:
        return None


get_fun = """
def get_script_str():
    return globals()['script_str']
"""


@wrapper_save_sha256
def serialize_to_json(obj, json_path, logger):
    """serialize obj to json and encrypt.
    :param logger:log handle
    :param obj: the object you want to serialize
    :param json_path: the path of json
    """
    try:
        logger.debug(f"delete old json in scripts.")
        if os.path.exists(json_path):
            os.chmod(json_path, stat.S_IWRITE)
            os.remove(json_path)
        with open(json_path, 'w') as wf:
            json.dump(obj, wf, default=default_f, sort_keys=False, indent=4)
            filename = json_path.replace('.json', '.py')
            if gv.IsHide:
                script_str = json.dumps(obj, default=default_f, sort_keys=False, indent=4)
                with open(filename, 'w') as f:
                    f.write(f'script_str = """\n{script_str}\n"""\n\n{get_fun}')
            else:
                if os.path.exists(filename):
                    os.remove(filename)
        logger.debug(f"serializeToJson success! {json_path}.")
    except Exception as e:
        logger.fatal(f'{currentframe().f_code.co_name}:{e}')
        QMessageBox.critical(None, 'Exception!', f'{currentframe().f_code.co_name}:{e}', QMessageBox.Yes)
        sys.exit(e)
