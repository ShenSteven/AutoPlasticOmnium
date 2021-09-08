#!/usr/bin/env python
# coding: utf-8
"""
@File   : testcase.py
@Author : Steven.Shen
@Date   : 2021/9/2
@Desc   : 
"""
from datetime import datetime
import json
import sys
from openpyxl import load_workbook
from bin.basefunc import IsNullOrEmpty, binary_read, get_sha256
from .suite import TestSuite
from .step import TestStep
from bin.globalconf import logger
from bin import globalvar as gv
from bin.globalvar import get_globalVal


class TestCase:
    test_station = ""
    _testcase_path_excel = None
    _testcase_path_json = None
    test_suites = []
    _tResult = True
    start_time = ""
    finish_time = ""

    def load_testcase_from_excel(self, temp_suite=None, workbook=None):
        suites_list = []
        itemHeader = []
        temp_suite_name = ""
        try:
            workbook = load_workbook(self._testcase_path_excel, read_only=True)
            worksheet = workbook[self.test_station]

            for i in list(worksheet.rows)[0]:  # 获取表头，第一行
                itemHeader.append(i.value)
                if not hasattr(TestStep, i.value):  # 动态创建Item类的属性
                    setattr(TestStep, i.value, '')
            for i in range(1, worksheet.max_row):  # 一行行的读取excel
                line = list(worksheet.rows)[i]
                if IsNullOrEmpty(line[1].value):  # ItemName为空停止解析
                    break
                if not IsNullOrEmpty(line[0].value):  # 新的seqItem
                    temp_suite_name = line[0].value
                    temp_suite = TestSuite(line[0].value, len(suites_list))
                    suites_list.append(temp_suite)

                test_step = TestStep()
                for header, cell in dict(zip(itemHeader, line)).items():  # 给step对象属性赋值
                    test_step.index = temp_suite.totalNumber
                    setattr(test_step, header, '' if IsNullOrEmpty(cell.value) else str(cell.value))
                    test_step.suite_name = temp_suite_name
                temp_suite.totalNumber = temp_suite.totalNumber + 1
                temp_suite.test_steps.append(test_step)
        except Exception as e:
            logger.exception(f"load testcase fail！！{e}")
            sys.exit(1)
        else:
            json.dump(suites_list, open(fr'F:\pyside2\scripts\{self.test_station}.json', 'w'),
                      default=lambda o: o.__dict__,
                      sort_keys=True,
                      indent=4)
            logger.info("load testcase success!")
            return suites_list
        finally:
            workbook.close()

    def load_sequences_from_json(self, shaPath, testcase_path_json):
        try:
            sha = binary_read(shaPath)
            logger.debug(f" txtSHA:{sha}")
            JsonSHA = get_sha256(self._testcase_path_json)
            logger.debug(f"jsonSHA:{JsonSHA}")
            if sha == JsonSHA:
                return json.load(open(testcase_path_json, 'r'))
            else:
                raise Exception(f"ERROR,json testCase file {testcase_path_json} has been tampered!")
        except Exception as e:
            logger.exception(e)
            sys.exit(1)

    def __init__(self, testcase_path, test_station):
        self.start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.test_station = test_station
        self._testcase_path_excel = testcase_path
        self.test_suites = self.load_testcase_from_excel(self._testcase_path_excel, test_station)

    def upload_result_to_mes(self, url):
        setattr(gv.mesPhases, 'status', str(self._tResult).upper())
        setattr(gv.mesPhases, 'start_time', self.start_time)
        setattr(gv.mesPhases, 'finish_time', self.finish_time)
        setattr(gv.mesPhases, 'error_code', get_globalVal('error_code_first_fail'))
        setattr(gv.mesPhases, 'error_details', get_globalVal('error_details_first_fail'))
        logger.debug(json.dumps(gv.mesPhases, default=lambda o: o.__dict__,
                                sort_keys=True,
                                indent=4))
        # response = requests.post(url, logger.debug(json.dumps(mesPhases, default=lambda o: o.__dict__,
        #                                                       sort_keys=True,
        #                                                       indent=4)))
        # if response.status_code == 200:
        #     return True
        # else:
        #     logger.debug(f'post fail:{response.content}')
        #     return False
        return True

    def run_suites(self, test_suites, global_fail_continue=False, suite_no=None):
        suite_result_list = []
        try:
            for suite in test_suites:
                if not IsNullOrEmpty(suite_no) and suite.index == suite_no:
                    pass
                else:
                    suite_result = suite.run_suite(global_fail_continue)
                    suite_result_list.append(suite_result)
                    if not suite_result and not global_fail_continue:
                        self._tResult = False
                        break
                    else:
                        pass
            self._tResult = all(suite_result_list)
            self.finish_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            suite_result_list.append(self.upload_result_to_mes(gv.mes_result))
            self._tResult = all(suite_result_list)

        except Exception as e:
            logger.exception(f"run_suites Exception！！{e}")
            self._tResult = False
            return self._tResult
        else:
            return self._tResult
        finally:
            ElapsedTime = datetime.strptime(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '%Y-%m-%d %H:%M:%S') \
                          - datetime.strptime(self.start_time, '%Y-%m-%d %H:%M:%S')
            logger.info(
                f"Test end and {'PASS' if self._tResult else 'FAIL'}!,ElapsedTime:{ElapsedTime},Symptom:"
                f"{get_globalVal('error_code_first_fail')}:{get_globalVal('error_details_first_fail')}.")
            logger.debug(json.dumps(gv.station, default=lambda o: o.__dict__,
                                    sort_keys=False,
                                    indent=4))


if __name__ == "__main__":
    # ss = load_sequences_from_excel("./Config/fireflyALL.xlsx", 'MBLT')  # ./Config/fireflyALL.xlsx
    testcase1 = TestCase(r"F:\pyside2\conf\fireflyALL.xlsx", 'MBLT')
    testingSuite = testcase1.test_suites.copy()
    testcase1.run_suites(testingSuite)
