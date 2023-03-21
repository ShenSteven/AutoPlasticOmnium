#!/usr/bin/env python
# coding: utf-8
"""
@File   : excel_style.py
@Author : Steven.Shen
@Date   : 3/10/2023
@Desc   : 
"""
import re

# font_sizes.py
import openpyxl
from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# def font_demo(path):
#     workbook = openpyxl.Workbook()
#     ws = workbook.active
#
#     data = [
#         ["Fruit", "Quantity"],
#         ["Kiwi", 3],
#         ["Grape", 15],
#         ["Apple", 3],
#         ["Peach", 3],
#         ["Pomegranate", 3],
#         ["Pear", 3],
#         ["Tangerine", 3],
#         ["Blueberry", 3],
#         ["Mango", 3],
#         ["Watermelon", 3],
#         ["Blackberry", 3],
#         ["Orange", 3],
#         ["Raspberry", 3],
#         ["Banana", 3]
#     ]
#
#     # cell = sheet["A1"]
#     # cell.font = Font(size=12)
#     # sheet["A1"].style = "Title"
#     # cell.value = "Hello"
#     # cell2 = sheet["A2"]
#     # cell2.font = Font(name="Arial", size=14, color="00FF0000")
#     # sheet["A2"] = "from"
#     # cell2 = sheet["A3"]
#     # cell2.font = Font(name="Tahoma", size=16, color="00339966")
#     # sheet["A3"] = "OpenPyXL"
#
#     from openpyxl.styles import NamedStyle, Font, Border, Side
#     highlight = NamedStyle(name="highlight")
#     highlight.font = Font(bold=True, size=20)
#     bd = Side(style='thick', color="000000")
#     highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#
#     for r in data:
#         ws.append(r)
#     d = ws.cell(row=ws.max_row, column=ws.max_column)
#     ws.auto_filter.ref = f"A1:{d.coordinate}"
#     for row in ws.rows:
#         for cell in row:
#             if cell.row % 2 == 0:
#                 print('0')
#             else:
#                 if cell.row == 1:
#                     cell.style = "Headline 4"
#                 print('1')
#     # ws[0].style = "Headline 4"
#     # for cell in ws[0]:
#     #     cell.style = "Headline 4"
#
#     # sheet.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
#     # print(sheet.auto_filter.filterColumn)
#
#     # sheet.auto_filter.add_sort_condition("B2:B15")
#     workbook.save(path)


def column_width_autofit(ws):
    # 设置一个字典用于保存列宽数据
    dims = {}
    # 遍历表格数据，获取自适应列宽数据
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        ws.column_dimensions[get_column_letter(col)].width = value + 2


from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()
ws = wb.active
for table in ws.tables.items():
    try:
        del ws.tables[table[0]]
    except KeyError:
        pass
ws.delete_rows(idx=1, amount=10 * 3)
data = [
    ['Apples', 10000, 5000, 8000, 8000],
    ['Pears', 2000, 3000, 4000, 5000],
    ['Bananas', 6000, '你好啊拉大手榴弹', 6500, 'fsdfs600000000fsfd'],
    ['Oranges', 500, 300, 200, 700],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
'''
# print(ws.tables)
ws.add_table(tab)
# hashed_password = ...
# print(wb.security)
# wb.security = WorkbookProtection(workbookPassword='0000', revisionsPassword='0000', lockWindows=True,
#                                  lockStructure=True, lockRevision=True)
# wb.security.set_workbook_password(hashed_password, already_hashed=True)
# ws.protection.sheet = True
# # ws.protection.enable()
# # ws.protection.disable()
# hashed_password = '...'
# ws.protection.password = hashed_password
ws.views.sheetView[0].zoomScale = 80
ssss(ws)

wb.save("table.xlsx")

if __name__ == "__main__":
    pass
    # font_demo("font_demo.xlsx")
