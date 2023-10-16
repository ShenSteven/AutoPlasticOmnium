#!/usr/bin/env python
# coding: utf-8
"""
@File   : basic func.py
@Author : Steven.Shen
@Date   : 2021/9/2
@Desc   : 
"""
import os
import csv
import platform
import re
import subprocess
import time
from collections import defaultdict
from datetime import datetime
from inspect import currentframe
from itertools import groupby
import openpyxl
import psutil
import hashlib
from socket import AddressFamily
from decimal import Decimal, ROUND_UP, ROUND_HALF_UP


# from pydub import AudioSegment

def ReplaceCommonEscapeSequences(date):
    return date.replace('\\n', '\n').replace('\\r', '\r')


def InsertCommonEscapeSequences(date):
    return date.replace('\n', '\\n').replace('\r', '\\r')


def bytes_to_string(byte_strs):
    str_list = []
    for i in range(len(byte_strs)):
        str_list.append('{:02X}'.format(byte_strs[i]))
    return str.join(' ', str_list)


def right_round(num, keep_n):
    """
    保留小数位，或者实现四舍五入
    :param num:
    :param keep_n:
    :return:
    """
    if isinstance(num, float):
        num = str(num)
    return Decimal(num).quantize((Decimal('0.' + '0' * keep_n)), rounding=ROUND_UP)


def CompareLimit(limitMin, limitMax, value, item, is_round=False):
    """比较测试limit"""
    if IsNullOrEmpty(limitMin) and IsNullOrEmpty(limitMax):
        return True, ''
    if IsNullOrEmpty(value):
        return False, ''
    temp = round(float(value)) if is_round else float(value)
    if IsNullOrEmpty(limitMin) and not IsNullOrEmpty(limitMax):  # 只需比较最大值
        item.logger.debug("compare Limit_max...")
        return temp <= float(limitMax), ''
    if not IsNullOrEmpty(limitMin) and IsNullOrEmpty(limitMax):  # 只需比较最小值
        item.logger.debug("compare Limit_min...")
        return temp >= float(limitMin), ''
    if not IsNullOrEmpty(limitMin) and not IsNullOrEmpty(limitMax):  # 比较最小最大值
        item.logger.debug("compare Limit_min and Limit_max...")
        if float(limitMin) <= temp <= float(limitMax):
            return True, ''
        else:
            if temp < float(limitMin):
                return False, 'TooLow'
            else:
                return False, 'TooHigh'


def subStr(SubStr1, SubStr2, revStr, item):
    """截取字符串"""
    if IsNullOrEmpty(SubStr1) and IsNullOrEmpty(SubStr2):
        return None
    elif IsNullOrEmpty(SubStr1):
        values = re.findall(f'^(.*?){SubStr2}', revStr)
    elif IsNullOrEmpty(SubStr2):
        values = re.findall(f'{SubStr1}(.*?)$', revStr)
    else:
        values = re.findall(f'{SubStr1}(.*?){SubStr2}', revStr)
    if len(values) == 1 and values[0] != '':
        testValue = values[0]
        item.logger.debug(f'get TestValue:{testValue}')
        return testValue.strip()
    else:
        raise Exception(f'get TestValue exception:{values}')


def assert_value(compInfo, item, rReturn):
    """判断测试结果"""
    if not IsNullOrEmpty(item.SPEC):
        try:
            rReturn = True if item.testValue in item.SPEC else False
        except TypeError as e:
            item.logger.error(f'{currentframe().f_code.co_name}:{e}')
    elif not IsNullOrEmpty(item.USL) or not IsNullOrEmpty(item.LSL):
        try:
            rReturn, compInfo = CompareLimit(item.LSL, item.USL, item.testValue, item)
        except TypeError as e:
            item.logger.error(f'{currentframe().f_code.co_name}:{e}')
    elif IsNullOrEmpty(item.USL) and IsNullOrEmpty(item.LSL):
        pass
    else:
        item.logger.warning(f"assert is unknown,SPEC:{item.SPEC},LSL:{item.LSL}USL:{item.USL}.")
    return compInfo, rReturn


def ensure_path_sep(path):
    """兼容 windows 和 linux 不同环境的操作系统路径 """
    if "/" in path:
        path = os.sep.join(path.split("/"))
    if "\\" in path:
        path = os.sep.join(path.split("\\"))
    return path


def ping(logger, host, timeout=1):
    """
    Returns True if host (str) responds to a ping request.
    Remember that a host may not respond to a ping (ICMP) request even if the host test_name is valid.
    """
    IsWind = platform.system() == 'Windows'
    param = '-n' if IsWind else '-cf'
    command = f'ping {param} 1 {host}'
    try:
        ret = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                             encoding=("gbk" if IsWind else "utf8"), timeout=timeout)
        if ret.returncode == 0 and 'TTL=' in ret.stdout:
            logger.debug(ret.stdout)
            return True
        else:
            logger.error(f"error:{ret.stdout},{ret.stderr}")
            return False
    except subprocess.TimeoutExpired:
        logger.debug(f'ping {host} Timeout.')
        raise
    except Exception as e:
        logger.fatal(e)
        raise


def run_cmd(logger, command, timeout=3):
    """send command, command executed successfully return true,otherwise false"""
    try:
        IsWind = platform.system() == 'Windows'
        logger.debug(f'run_cmd-->{command}')
        ret = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                             encoding=("gbk" if IsWind else "utf8"), timeout=timeout, input='')
        if ret.returncode == 0:  # 表示命令下发成功，不对命令内容结果做判断
            logger.debug(ret.stdout)
            return True, ret.stdout
        else:
            logger.error(f"error:{ret.stderr}")
            return False, ret.stdout
    except Exception as e:
        logger.fatal(e)
        raise


def kill_process(logger, process_name, killall=True):
    try:
        for pid in psutil.pids():
            if psutil.pid_exists(pid):
                p = psutil.Process(pid)
                if p.name() == process_name:
                    p.kill()
                    logger.debug(f"kill pid-{pid},test_name-{p.name()}")
                    time.sleep(1)
                    if not killall:
                        break
        return True
    except Exception as e:
        logger.fatal(e)
        raise e


def process_exists(process_name):
    pids = psutil.pids()
    ps = [psutil.Process(pid) for pid in pids]
    process_names = [p.name for p in ps]
    if process_name in process_names:
        return True
    else:
        return False


def start_process(logger, full_path, process_name):
    """if process exists, return , otherwise start it and check"""
    try:
        if not process_exists(process_name):
            run_cmd(logger, full_path)
            time.sleep(3)
            return process_exists(process_name)
        else:
            return True
    except Exception as e:
        logger.fatal(e)
        raise e
        # return False


def restart_process(logger, full_path, process_name):
    """kill and start"""
    try:
        if kill_process(logger, process_name):
            return start_process(logger, full_path, process_name)
    except Exception as e:
        logger.fatal(e)
        raise e
        # return False


# def save_config(logger, yaml_file, obj):
#     with open(yaml_file, mode='r+', encoding='utf-8') as f:
#         readall = f.read()
#         try:
#             js = json.dumps(obj, default=lambda o: o.__dict__, sort_keys=False, indent=4)
#             ya = yaml.safe_load(js)
#             f.seek(0)
#             yaml.safe_dump(ya, stream=f, default_flow_style=False, sort_keys=False, indent=4)
#         except Exception as e:
#             logger.fatal(f"save_config! {e}")
#             f.seek(0)
#             f.write(readall)
#             raise
#         else:
#             logger.info(f"save conf.yaml success!")


def wrapper(logger, flag):
    def wrapper_inner(func):
        def inners(**kwargs):
            start = None
            if flag:
                start = datetime.now()
            ret = func(**kwargs)
            if flag:
                logger.debug(f'elapsed time:{datetime.now() - start}')
            return ret

        return inners

    return wrapper_inner


def wrapper_time(logger, fun):
    def inner(*args):
        start = datetime.now()
        fun(*args)
        logger.debug(f'elapsed time:{datetime.now() - start}')

    return inner


def create_csv_file(logger, filename, header, updateColumn=False):
    try:
        if not os.path.exists(filename):
            with open(filename, 'w', newline='') as f:
                file = csv.writer(f)
                file.writerow(header)
            logger.debug(f'create_csv_file:{filename}')
        else:
            if updateColumn:
                with open(filename, 'r+', newline='') as rwf:
                    reader = csv.reader(rwf)
                    rows = [row for row in reader]
                    if [0] != header:
                        rows[0] = header
                        rwf.seek(0)
                        file = csv.writer(rwf)
                        file.writerows(rows)
                    logger.debug(f'update csvHeader success!')
    except Exception as e:
        logger.fatal(e)
        raise e


def write_csv_file(logger, filename, row):
    try:
        with open(filename, 'a', newline='') as f:
            file = csv.writer(f)
            file.writerow(row)
    except Exception as e:
        logger.fatal(e)
        raise e


def IsNullOrEmpty(strObj: str):
    if strObj and len(str(strObj)) > 0:
        return False
    else:
        return True


def binary_read(filepath):
    with open(filepath, 'rb') as f:
        content = f.read()
        return content.decode(encoding='UTF-8')


def binary_write(filepath, content):
    with open(filepath, 'wb') as f:
        f.write(content.encode('utf8'))


def get_sha256(filepath) -> str:
    with open(filepath, 'rb') as f:
        sha256obj = hashlib.sha256()
        sha256obj.update(f.read())
        return sha256obj.hexdigest()


def GetAllIpv4Address(networkSegment):
    for name, info in psutil.net_if_addrs().items():
        for addr in info:
            if AddressFamily.AF_INET == addr.family and str(addr.address).startswith(networkSegment):
                return str(addr.address)


def get_file_ext_list(path_dir, ext):
    """
    Gets the files in the directory with the specified suffix.
    :param path_dir:
    :param ext: .exe, .py and so on.
    :return: list
    """
    L = []
    for root, dirs, files in os.walk(path_dir):
        for file in files:
            if os.path.splitext(file)[1] == ext:
                L.append(os.path.join(root, file))
    return L


# def audio_play(audio_path):
#     """use ffmpeg play audio"""
#
#     def thread_update():
#         song = AudioSegment.from_wav(audio_path)
#         play(song)
#
#     thread = Thread(target=thread_update, daemon=True)
#     thread.start()


# def play_audio(path, parent=None):
#     def thread_update():
#         player = QMediaPlayer(parent)
#         qurl = QUrl.fromLocalFile(path)
#         qmusic = QMediaContent(qurl)
#         player.setMedia(qmusic)
#         player.setVolume(100)
#         player.play()
#
#     thread = Thread(target=thread_update, daemon=True)
#     thread.start()


# def audio_to_export_30s(sourcePath, wavePath, start):
#     wav = AudioSegment.from_wav(sourcePath)
#     wav[start * 1000:(start + 30) * 1000].export(wavePath, format="wav")


def str_to_int(strs):
    try:
        num = int(strs)
        return True, num
    except:
        return False, 0


def fetch_logo_data(filename, sheet_name, row_start, row_end, column_start, column_end):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[sheet_name]
    cell_data = []

    mark_color_list = []
    for r in range(10, 0, -1):
        mark_color_value = worksheet.cell(row=4, column=r + 3).fill.start_color.rgb
        mark_color_list.append(mark_color_value)
    print(mark_color_list)

    for r in range(row_start, row_end + 1):
        row_color = []
        for col in range(column_start, column_end + 1):
            cell_color = worksheet.cell(row=r, column=col).fill.start_color.rgb
            if cell_color in mark_color_list:
                rgbColor = mark_color_list.index(cell_color) + 1
            else:
                rgbColor = 0
            row_color.append(rgbColor)
        cell_data.append(tuple(row_color))
    workbook.close()
    return cell_data


def de_adjacent_repeat(my_list: list or tuple):
    """Remove adjacent duplicates from the list, and get index range list"""
    nodup_list = [k for k, group in groupby(my_list)]  # 去除相邻重复项
    index_list = []
    index = 0
    while index < len(my_list):
        start_position = index
        val = my_list[index]
        while index < len(my_list) and my_list[index] == val:
            index += 1
        end_position = index - 1
        index_list.append([start_position, end_position])
    print(nodup_list)
    print(index_list)
    return nodup_list, index_list


def get_nodup_list(my_list: list, sort=True) -> list:
    """
    Remove duplicates from the list.
    :param sort: keep the original order
    :param my_list:
    :return:
    """
    new_li = list(set(my_list))
    if sort:
        new_li.sort(key=my_list.index)
    print(new_li)
    return new_li


def merge_list_to_dict(list_one: list or tuple, list_two: list or tuple) -> dict:
    """
    merge two lists to a dict, if list_one have duplicate key, append to a list value
    :param list_one:
    :param list_two:
    :return:
    """
    d = defaultdict(list)
    for key, value in zip(list_one, list_two):
        d[key].append(value)
    print(dict(d))
    return dict(d)


if __name__ == '__main__':
    pass
    print(right_round(4.94524324, 2))
    # li = [('a', 'a'), ('a', 'a'), ('a', 'a'), ('bb', 'bb'), ('a', 'a'), ('c', 'c'), ('c', 'c')]
    # print(li)
    # get_nodup_list(li)
    # r1, r2 = de_adjacent_repeat(li)
    # dd = merge_list_to_dict(r1, r2)
    # print(dd.keys())
    # print(dd.values())
    # create_csv_file('test.csv', ['No', 'Phase test_name', 'Test test_name', 'Error Code'])
    # write_csv_file('test.csv', ['1wqw', '2', '3', '4'])
    # print(str(True))
    # ping('192.168.1.101')
    # ping('127.0.0.1')
    # aa = restart_process('EXCEL.EXE')
    # logger.debug(f"ds{aa}")
    # register("demon", "1@1.com")  # test_name:%s, age:%s, others:%s ('demon', '1@1.com', {})
    # register("demon", "1@1.com",
    #          addr="shanghai")  # test_name:%s, age:%s, others:%s ('demon', '1@1.com', {'addr': 'shanghai'})
